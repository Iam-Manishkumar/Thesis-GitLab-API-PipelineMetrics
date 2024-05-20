import requests
from openpyxl import Workbook

# GitLab API URL and access token
api_url = "enter your text"
access_token = "enter your text"

# Define project IDs and their corresponding stage names
project_stage_names = {
    enter your text: ["preview", "deploy", "configuration", "cleanup-deployment", "destroy"],
    enter your text: ["plan", "apply", "configuration", "cleanup-deployment", "cleanup"],
    enter your text: ["sonarqube-check", "sonarqube-vulnerability-report", "Unit-Testing", "FS-NVD-Scan", "Build", "Multi-Project", "Push-based"]
}

# Function to format duration
def format_duration(minutes, seconds):
    return f"{minutes:02d}:{int(seconds):02d}"

# Function to fetch pipeline duration, status, and stage durations
def get_pipeline_info(project_id, pipeline_id):
    headers = {"PRIVATE-TOKEN": access_token}
    url = f"{api_url}projects/{project_id}/pipelines/{pipeline_id}/jobs"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        jobs_data = response.json()
        stages_data = {}
        for job in jobs_data:
            stage = job["stage"]
            duration = job["duration"]
            if duration is not None and stage in project_stage_names[project_id]:
                stages_data[stage] = stages_data.get(stage, 0) + duration
        
        # Calculate total duration of the pipeline
        total_duration_seconds = sum(stages_data.values())
        
        # Convert total duration to minutes and seconds
        total_minutes = int(total_duration_seconds / 60)
        total_seconds = total_duration_seconds % 60
        
        # Check if any job is still running or has a status other than "success"
        running_jobs = any(job["status"] == "running" for job in jobs_data)
        if running_jobs:
            status = "running"
        else:
            status = "failed" if any(job["status"] != "success" for job in jobs_data) else "success"
        
        return pipeline_id, project_id, status, stages_data, total_minutes, total_seconds
    else:
        print(f"Failed to fetch pipeline {pipeline_id}: {response.status_code}")
        return None

# Function to fetch all pipelines for a project
def get_project_pipelines(project_id, limit=60):
    headers = {"PRIVATE-TOKEN": access_token}
    url = f"{api_url}projects/{project_id}/pipelines?per_page={limit}"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        pipelines_data = response.json()
        pipelines = []
        for pipeline in pipelines_data:
            pipeline_id = pipeline["id"]
            pipeline_info = get_pipeline_info(project_id, pipeline_id)
            if pipeline_info:
                pipelines.append(pipeline_info)
        return pipelines
    else:
        print(f"Failed to fetch pipelines for project {project_id}: {response.status_code}")
        return None

# Create a new Excel workbook
wb = Workbook()

# Delete the default sheet created by Workbook
default_sheet = wb.active
wb.remove(default_sheet)

for project_id, stages in project_stage_names.items():
    # Fetch and write pipeline info to Excel for the latest 50 pipelines
    pipelines = get_project_pipelines(project_id)
    if pipelines:
        ws = wb.create_sheet(title=f"Project {project_id}")  # Create a new worksheet for each project
        headers = ["Pipeline ID", "Project ID", "Status"] + stages + ["Total Duration"]
        ws.append(headers)
        for pipeline in pipelines:
            pipeline_id, project_id, status, stages_data, total_minutes, total_seconds = pipeline
            row_data = [pipeline_id, project_id, status]
            for stage in stages:
                stage_duration = stages_data.get(stage, 0)
                stage_minutes = int(stage_duration / 60)
                stage_seconds = round(stage_duration % 60, 2)
                row_data.append(format_duration(stage_minutes, stage_seconds))
            row_data.append(format_duration(total_minutes, total_seconds))
            ws.append(row_data)

# Remove formatting from cells
for ws in wb:
    for row in ws.iter_rows():
        for cell in row:
            cell.number_format = 'General'

# Save the workbook
wb.save("pipeline_info.xlsx")
